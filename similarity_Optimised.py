import os
import pandas as pd
import datetime
from difflib import SequenceMatcher
import re
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
from pathlib import Path
from IPython.display import clear_output
from itertools import combinations
import matplotlib.pyplot as plt
import matplotlib.colors as mcolors
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font

df = pd.read_csv('R_with_patterns.csv')

def convert_bytes_to_gb(df, column_name):
    """
    Convert a column with sizes in bytes to a column with sizes in gigabytes (GB).

    Parameters:
    df (pd.DataFrame): The DataFrame containing the column to convert.
    column_name (str): The name of the column with sizes in bytes.

    Returns:
    pd.DataFrame: The DataFrame with an additional column for sizes in gigabytes.
    """
    # Check if the column exists in the DataFrame
    if column_name not in df.columns:
        raise ValueError(f"Column '{column_name}' does not exist in the DataFrame")

    # Convert the size from bytes to GB
    df[f'{column_name}_in_gb'] = (df[column_name] / (1024 ** 3)).round(2)

    return df

# def similar(a, b):
#     return SequenceMatcher(None, a, b).ratio()
def similar(a, b):
    if a==b:
        return True
    else:
        return False

def find_similar_values(df, column_name, threshold=0.8):
    final_data = []
    seen_pairs = set()  # Keep track of pairs we've already processed

    # Create a dictionary to store the values and their corresponding rows
    value_dict = {i: df.iloc[i] for i in range(len(df))}

    # Use combinations to generate unique pairs
    for i, j in combinations(range(len(df)), 2):
        original_row = value_dict[i]
        similar_row = value_dict[j]
        original_value = original_row[column_name]
        similar_value = similar_row[column_name]

        if not isinstance(original_value, str) or not isinstance(similar_value, str):
            continue

        # Create a frozen set of the pair to ensure order doesn't matter
        pair = frozenset([original_row['folder_name'], similar_row['folder_name']])

        # Only process if we haven't seen this pair before
        if pair not in seen_pairs and similar(original_value, similar_value) ==True:
            seen_pairs.add(pair)  # Mark this pair as seen
            final_data.append([original_value, original_row['folder_path'],
                               original_row['creation_date_str'], original_row['size_bytes_in_gb'],
                               original_row['num_matching_files'],original_row['num_pattern_files'],
                               similar_value, similar_row['folder_path'],
                               similar_row['creation_date_str'], similar_row['size_bytes_in_gb'],
                               similar_row['num_matching_files'],similar_row['num_pattern_files']])
            
def filter_and_sort_similar_results(similar_results_df):
    """
    Filters the DataFrame to include only rows where:
    - 'Folder Name (Original)' and 'Folder Name (Similar)' are the same.
    - 'Size, GB (Original)' and 'Size, GB (Similar)' are different by no more than 10%.
    Sorts the resulting DataFrame based on 'Size, GB (Original)' from highest to lowest.

    Args:
        similar_results_df (pd.DataFrame): The input DataFrame.

    Returns:
        pd.DataFrame: The filtered and sorted DataFrame.
    """
    # Filter rows where 'Folder Name (Original)' and 'Folder Name (Similar)' are the same
    filtered_df = similar_results_df[similar_results_df['Folder Name (Original)'] == similar_results_df['Folder Name (Similar)']]

    # Calculate the percentage difference in size
    size_diff = abs(filtered_df['Size, GB (Original)'] - filtered_df['Size, GB (Similar)']) / filtered_df['Size, GB (Similar)']

    # Filter rows where the size difference is no more than 10%
    updated_data = filtered_df[size_diff <= 0.10]

    # Sort the resulting DataFrame based on 'Size, GB (Original)' from highest to lowest
    # updated_data = updated_data.sort_values(by='Size, GB (Original)', ascending=False)
    updated_data = updated_data.sort_values(by='Folder Path (Original)', key=lambda x: x.str.len())

    return updated_data

updated_data = filter_and_sort_similar_results(similar_results_df)
updated_data.head()

updated_data = updated_data[
    (updated_data['Number of files ndpi_qptiff_tif_tiff (Original)'] > 0) &
    (updated_data['Number of files ndpi_qptiff_tif_tiff (Similar)'] > 0) 
]

updated_data.to_csv('updated_data_filtered.csv', index=False)


def highlight_differences(df, output_file):
    """
    Compares columns 'A' and 'C', 'B' and 'D' in a pandas DataFrame
    and highlights the differences in an Excel worksheet.

    Args:
        df: pandas DataFrame with columns 'A', 'B', 'C', 'D'.
        output_file: Path to the output Excel file.
    """
    aqua_fill = PatternFill(start_color="00FFFF", end_color="00FFFF", fill_type="solid")

    # Create a new workbook and select the active sheet
    wb = Workbook()
    ws = wb.active

    # Write DataFrame to the worksheet
    for r in dataframe_to_rows(df, index=True, header=True):
        ws.append(r)

    # Define your comments for each column
    comments = [ ' ','Yellow = not equal', 'Blue = not equal ', 'Purple if diffenece > 5 min', 'Green if size diff > 5%',
                'Yellow = not equal', 'Blue = not equal ', 'Purple if diffenece > 5 min', 'Green if size diff > 5%']
  # Add comments for each column

    # Insert a new row right after the header
    ws.insert_rows(2)  # Insert at row 2 (right after header)
    ws.delete_rows(3)
   

    # Add comments in the inserted row
    for col_idx, comment in enumerate(comments, start=1):  # start=1 if you have index column
        cell = ws.cell(row=2, column=col_idx)
        cell.value = comment
        cell.font = Font(italic=True, size=9, color="808080")  # Gray, italic, smaller font

    # Write the rest of the data
    for r_idx, row in enumerate(dataframe_to_rows(df, index=True, header=False), start=3):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    # Highlight differences in columns 'Folder name'
    for row in range(3, ws.max_row + 1):
        a_val = ws.cell(row=row, column=2).value
        c_val = ws.cell(row=row, column=8).value
        if a_val != c_val:
            # # Highlight if SequenceMatcher ratio is below a threshold
            # if SequenceMatcher(None, str(a_val), str(c_val)).ratio() >= 0.9:  # Adjust threshold as needed
                ws.cell(row=row, column=2).fill = PatternFill(start_color="FFFFE040", end_color="FFFFE040", fill_type="solid")
                ws.cell(row=row, column=8).fill = PatternFill(start_color="FFFFE040", end_color="FFFFE040", fill_type="solid")

    # Highlight differences in columns 'Path'
    for row in range(3, ws.max_row + 1):
        b_val = ws.cell(row=row, column=3).value
        d_val = ws.cell(row=row, column=9).value
        if b_val != d_val:
            # # Highlight if SequenceMatcher ratio is below a threshold
            # if SequenceMatcher(None, str(b_val), str(d_val)).ratio() >= 0.9:  # Adjust threshold as needed
                ws.cell(row=row, column=3).fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")  # Blue fill
                ws.cell(row=row, column=9).fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")  # Blue fill

    
    # Highlight differences in columns 'Size'
    for row in range(3, ws.max_row + 1):
        e_val = ws.cell(row=row, column=5).value  # Column 'E'
        f_val = ws.cell(row=row, column=11).value  # Column 'F'

        # if isinstance(e_val, int) and isinstance(f_val, int):
        # Calculate percentage difference
        
        if f_val != 0:  # Prevent division by zero
            percent_diff = abs((e_val - f_val) / f_val) * 100
                
                # Check if difference is within 5%
            if percent_diff >= 5:
                ws.cell(row=row, column=5).fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                ws.cell(row=row, column=11).fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")

     
     # Highlight differences in columns 'Creation date'
    for row in range(3, ws.max_row + 1):
        g_val = ws.cell(row=row, column=4).value  # Column 'G'
        h_val = ws.cell(row=row, column=10).value  # Column 'H'

        # Try to convert to datetime objects, handle potential errors
        try:
            g_val = datetime.strptime(g_val, '%Y-%m-%d %H:%M:%S')
            h_val = datetime.strptime(h_val, '%Y-%m-%d %H:%M:%S')
        except (TypeError, ValueError):
            # If conversion fails, skip highlighting for this row
            continue

        time_difference = abs((g_val - h_val).total_seconds() / 60)  # Convert to minutes
        if time_difference > 5:
            ws.cell(row=row, column=4).fill = PatternFill(start_color="E3D6E6", end_color="E3D6E6", fill_type="solid")  # Light purple fill
            ws.cell(row=row, column=10).fill = PatternFill(start_color="E3D6E6", end_color="E3D6E6", fill_type="solid")  # Light purple fill
    
    wb.save(output_file)